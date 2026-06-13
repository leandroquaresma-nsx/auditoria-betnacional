import os
import re
import base64
import html
import traceback
import pandas as pd
import openpyxl
import matplotlib
matplotlib.use('Agg') 
import matplotlib.pyplot as plt
from datetime import datetime

from flask import Flask, request, jsonify
from flask_cors import CORS
from google import genai

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

app = Flask(__name__)
CORS(app) 

CHAVE_API = os.environ.get("GEMINI_API_KEY")
client = genai.Client(api_key=CHAVE_API)

LOGO_PATH = "logo.png"
SLA_META_MINUTOS = 15
CUSTO_HORA_ATENDENTE = 18.50 

def limpar_dados_sensiveis(texto):
    if not isinstance(texto, str): return texto
    texto = re.sub(r'[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+', '[EMAIL PROTEGIDO]', texto)
    texto = re.sub(r'\b\d{3}\.?\d{3}\.?\d{3}-?\d{2}\b', '[CPF PROTEGIDO]', texto)
    texto = re.sub(r'\(?\d{2}\)?\s?\d{4,5}-?\d{4}', '[TELEFONE PROTEGIDO]', texto)
    return texto

def cacar_nome_atendente(conversa):
    if not isinstance(conversa, str): return "Não identificado"
    
    match = re.search(r'\b([A-Za-zÀ-ÿ]+(?: [A-Za-zÀ-ÿ]+)* [A-Za-zÀ-ÿ]\.)', conversa)
    if match: return match.group(1).title()
        
    falantes = re.findall(r'\(\d{2}:\d{2}:\d{2}\)\s+([^:]+):', conversa)
    termos_robo = ['betnacional', 'bot', 'suporte', 'atendimento', 'web user']
    
    for falante in falantes:
        f_limpo = falante.strip()
        f_lower = f_limpo.lower()
        if not any(robo in f_lower for robo in termos_robo):
            if re.search(r'\b[A-Za-zÀ-ÿ]\.?$', f_limpo):
                return f_limpo.title()
                
    return "Não identificado"

def classificar_origem(motivo):
    motivo = str(motivo).lower()
    if any(k in motivo for k in ['senha', 'acesso', 'dados', 'mfa', 'usuário', 'esqueci', 'tutorial']):
        return 'Erro/Dúvida do Cliente'
    return 'Falha da Plataforma/Sistema'

def calcular_tma(conversa):
    if not isinstance(conversa, str): return 0
    tempos = re.findall(r'\[(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\]', conversa)
    if len(tempos) >= 2:
        try:
            t_inicio = datetime.strptime(tempos[0], '%Y-%m-%d %H:%M:%S')
            t_fim = datetime.strptime(tempos[-1], '%Y-%m-%d %H:%M:%S')
            minutos = abs((t_fim - t_inicio).total_seconds() / 60.0)
            return minutos if minutos < 300 else 0
        except: return 0
    return 0

def extrair_hora(conversa):
    if not isinstance(conversa, str): return None
    tempos = re.findall(r'\[\d{4}-\d{2}-\d{2} (\d{2}):\d{2}:\d{2}\]', conversa)
    if tempos: return int(tempos[0])
    return None

def detetar_risco_churn(conversa):
    if not isinstance(conversa, str): return False
    conversa_limpa = conversa.lower()
    termos_risco = ['cancelar', 'procon', 'reclame aqui', 'advogado', 'processar', 'roubo', 'fraude', 'lixo', 'encerrar conta', 'devolver meu dinheiro']
    return any(termo in conversa_limpa for termo in termos_risco)

def adicionar_rodape(canvas, doc):
    canvas.saveState()
    canvas.setFont('Helvetica', 8)
    canvas.setFillColor(colors.HexColor('#64748B'))
    canvas.drawString(35, 20, "Betnacional | Data Science & Predictive Analytics Report - CONFIDENCIAL")
    canvas.drawRightString(letter[0]-35, 20, f"Página {doc.page}")
    canvas.restoreState()

# === ROTA RAIZ ADICIONADA PARA RESOLVER O ERRO 404 DO RENDER ===
@app.route("/", methods=["GET"])
def home():
    return "<h1>Servidor da Betnacional 100% Operacional 🚀</h1><p>Aguardando planilhas via POST na rota /auditar</p>", 200

@app.route("/auditar", methods=["POST"])
def auditar():
    try:
        if "file" not in request.files:
            return jsonify({"status": "erro", "mensagem": "Arquivo ausente."}), 400
            
        file = request.files["file"]
        df = pd.read_csv(file) if file.filename.endswith('.csv') else pd.read_excel(file)
        
        total_geral_casos = len(df)
        
        df['atendente_extraido'] = df['comments'].apply(cacar_nome_atendente)
        df['origem_erro'] = df['reason'].apply(classificar_origem)
        df['tags'] = df['tags'].fillna('')
        df['is_inativo'] = df['tags'].str.contains('inativo', case=False)
        df['tma_minutos'] = df['comments'].apply(calcular_tma)
        df['cumpriu_sla'] = df['tma_minutos'] <= SLA_META_MINUTOS
        df['hora_contato'] = df['comments'].apply(extrair_hora)
        df['fcr_sucesso'] = (df['tma_minutos'] > 0) & (df['tma_minutos'] <= 10) & (~df['is_inativo'])
        df['risco_churn_critico'] = df['comments'].apply(detetar_risco_churn) 
        
        for col in ['comments', 'ticket_summary']:
            if col in df.columns: df[col] = df[col].apply(limpar_dados_sensiveis)

        qtd_inativos = int(df['is_inativo'].sum())
        pct_inativos = (qtd_inativos / total_geral_casos) * 100
        pct_fcr = (df['fcr_sucesso'].sum() / total_geral_casos) * 100
        qtd_clientes_risco = int(df['risco_churn_critico'].sum())
        
        origem_counts = df['origem_erro'].value_counts()
        pct_erro_cliente = (int(origem_counts.get('Erro/Dúvida do Cliente', 0)) / total_geral_casos) * 100
        
        casos_validos_tma = df[df['tma_minutos'] > 0]
        tma_medio_global = casos_validos_tma['tma_minutos'].mean() if not casos_validos_tma.empty else 0
        pct_dentro_sla = (len(casos_validos_tma[casos_validos_tma['cumpriu_sla']]) / len(casos_validos_tma) * 100) if not casos_validos_tma.empty else 0

        prejuizo_estimado = ((tma_medio_global * total_geral_casos) / 60) * CUSTO_HORA_ATENDENTE

        df['custo_atendimento'] = (df['tma_minutos'] / 60) * CUSTO_HORA_ATENDENTE
        df_agentes = df[df['atendente_extraido'] != "Não identificado"].groupby('atendente_extraido').agg(
            Volume=('reason', 'count'), Custo_Gerado=('custo_atendimento', 'sum')
        ).sort_values(by='Custo_Gerado', ascending=False).head(5)

        motivos_agrupados = df.groupby('reason').agg(Quantidade=('reason', 'count')).reset_index().sort_values(by='Quantidade', ascending=False)
        motivos_top5 = motivos_agrupados.head(5)
        top_motivo_nome = str(motivos_top5.iloc[0]['reason']).replace('_', ' ').title() if not motivos_top5.empty else "N/A"

        volume_hora = df['hora_contato'].dropna().value_counts().sort_index()
        hora_pico = int(volume_hora.idxmax()) if not volume_hora.empty else 0

        # --- GERADOR DE GRÁFICOS ---
        cores_bet = ['#FF5A00', '#1C2541', '#3A506B', '#5BC0BE', '#CBD5E1']
        font_titulo = {'fontsize': 10, 'fontweight': 'bold', 'color': '#0B132B'}
        
        fig1, ax1 = plt.subplots(figsize=(6.5, 3))
        if not motivos_top5.empty:
            ax1.pie(motivos_top5['Quantidade'], colors=cores_bet, startangle=140, pctdistance=0.75, textprops=dict(color="white", weight="bold", fontsize=8))
            ax1.add_artist(plt.Circle((0,0), 0.55, fc='white'))
            ax1.legend([m.replace('_', ' ').title() for m in motivos_top5['reason']], loc="center left", bbox_to_anchor=(0.9, 0.5), frameon=False, fontsize=8)
        ax1.set_title('Concentração de Volume', fontdict=font_titulo, loc='left')
        plt.tight_layout()
        path_motivos = "/tmp/g_motivos.png"
        fig1.savefig(path_motivos, dpi=150)
        plt.close(fig1)

        fig2, ax2 = plt.subplots(figsize=(4, 3))
        ax2.pie([total_geral_casos - qtd_clientes_risco, qtd_clientes_risco], labels=['Normal', 'Ameaça / Churn'], autopct='%1.1f%%', colors=['#1C2541', '#EF4444'], startangle=90, textprops=dict(color="white", weight="bold", fontsize=9))
        ax2.set_title('Índice de Risco (Churn)', fontdict=font_titulo)
        plt.tight_layout()
        path_risco = "/tmp/g_risco.png"
        fig2.savefig(path_risco, dpi=150)
        plt.close(fig2)

        fig3, ax3 = plt.subplots(figsize=(4, 3))
        if not df_agentes.empty:
            ax3.barh(df_agentes.index[::-1], df_agentes['Custo_Gerado'][::-1], color='#FF5A00')
        ax3.set_title('Custo Operacional (R$) por Analista', fontdict=font_titulo, loc='left')
        ax3.spines['top'].set_visible(False)
        ax3.spines['right'].set_visible(False)
        plt.tight_layout()
        path_atendentes = "/tmp/g_atendentes.png"
        fig3.savefig(path_atendentes, dpi=150)
        plt.close(fig3)

        fig4, ax4 = plt.subplots(figsize=(6.5, 2.5))
        if not volume_hora.empty:
            ax4.plot(volume_hora.index, volume_hora.values, color='#3A506B', marker='o', linewidth=2)
            ax4.fill_between(volume_hora.index, volume_hora.values, color='#3A506B', alpha=0.1)
            ax4.set_xticks(range(0, 24, 2))
        ax4.set_title('Distribuição Horária (WFM)', fontdict=font_titulo, loc='left')
        ax4.grid(axis='y', linestyle='--', alpha=0.5)
        ax4.spines['top'].set_visible(False)
        ax4.spines['right'].set_visible(False)
        plt.tight_layout()
        path_horario = "/tmp/g_horario.png"
        fig4.savefig(path_horario, dpi=150)
        plt.close(fig4)

        # --- IA ---
        prompt = (f"Atue como Chief Data Officer (CDO) da Betnacional. Analise estes dados auditados pelo sistema de IA:\n"
                  f"- Ocorrências: {total_geral_casos} | SLA: {pct_dentro_sla:.1f}% | Prejuízo estimado: R$ {prejuizo_estimado:,.2f}.\n"
                  f"- ALERTA VERMELHO: {qtd_clientes_risco} clientes mencionaram intenção de cancelar ou processos no chat.\n"
                  f"- Gargalo Crítico: {top_motivo_nome} com {pct_inativos:.1f}% de abandono total.\n"
                  f"Crie um Laudo de Risco Preditivo com EXATAMENTE estes 4 tópicos (seja altamente formal):\n"
                  f"1. AUDITORIA FINANCEIRA (Descreva a perda de R$ {prejuizo_estimado:,.2f})\n"
                  f"2. RADAR DE CRISE E CHURN (Discuta a gravidade dos {qtd_clientes_risco} clientes em risco)\n"
                  f"3. PERFORMANCE E CUSTO HUMANO (Analise o impacto do TMA no custo do call center)\n"
                  f"4. PLANO DE CONTENÇÃO (3 estratégias urgentes para mitigar processos)")
        try:
            texto_ia = client.models.generate_content(model='gemini-2.5-flash', contents=prompt).text
        except:
            texto_ia = "Falha ao processar análise preditiva da IA."

        # --- EXCEL ---
        nome_excel = "/tmp/Relatorio_Auditoria_Betnacional.xlsx"
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = 'Painel Preditivo de Risco'
        ws1.views.sheetView[0].showGridLines = False
        
        fill_fundo = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        for row in ws1.iter_rows(min_row=1, max_row=50, min_col=1, max_col=15):
            for cell in row: cell.fill = fill_fundo

        borda_kpi = Border(left=Side(style='thin', color="CBD5E1"), right=Side(style='thin', color="CBD5E1"), top=Side(style='thin', color="CBD5E1"), bottom=Side(style='thin', color="CBD5E1"))
        
        linha_kpi = 2
        if os.path.exists(LOGO_PATH):
            img = openpyxl.drawing.image.Image(LOGO_PATH)
            img.width, img.height = 160, 40
            ws1.add_image(img, 'B2')
            linha_kpi = 5
        else:
            ws1['B2'] = "DATA SCIENCE & RISK MANAGEMENT DASHBOARD"
            ws1['B2'].font = Font(size=16, bold=True, color="0B132B")
            linha_kpi = 4
        
        kpis = [
            ("VOLUMETRIA TOTAL", total_geral_casos, f'B{linha_kpi}', f'B{linha_kpi+1}', "1C2541"),
            ("AMEAÇA LEGAL / CHURN", f"{qtd_clientes_risco} Clientes", f'C{linha_kpi}', f'C{linha_kpi+1}', "EF4444"),
            ("CUSTO (DESPERDÍCIO)", f"R$ {prejuizo_estimado:,.0f}", f'D{linha_kpi}', f'D{linha_kpi+1}', "EF4444"),
            ("SUCESSO SLA", f"{pct_dentro_sla:.1f}%", f'E{linha_kpi}', f'E{linha_kpi+1}', "1C2541"),
            ("FCR RÁPIDO", f"{pct_fcr:.1f}%", f'F{linha_kpi}', f'F{linha_kpi+1}', "10B981"),
            ("TAXA ABANDONO", f"{pct_inativos:.1f}%", f'G{linha_kpi}', f'G{linha_kpi+1}', "1C2541")
        ]
        
        for titulo, valor, c_tit, c_val, cor_texto in kpis:
            ws1[c_tit] = titulo
            ws1[c_tit].font = Font(size=8, bold=True, color="64748B")
            ws1[c_val] = valor
            ws1[c_val].font = Font(size=14, bold=True, color=cor_texto)
            ws1[c_val].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") 
            ws1[c_val].border = borda_kpi

        for col, width in zip(['A','B','C','D','E','F','G','H'], [2, 18, 25, 20, 18, 18, 18, 2]): ws1.column_dimensions[col].width = width

        ws1.add_image(openpyxl.drawing.image.Image(path_motivos), f'B{linha_kpi+3}')
        ws1.add_image(openpyxl.drawing.image.Image(path_risco), f'F{linha_kpi+3}')
        ws1.add_image(openpyxl.drawing.image.Image(path_horario), f'B{linha_kpi+19}')
        ws1.add_image(openpyxl.drawing.image.Image(path_atendentes), f'F{linha_kpi+19}')

        ws2 = wb.create_sheet('Database com Heatmap')
        ws2.append(list(df.columns))
        for cell in ws2[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = PatternFill(start_color="1C2541", end_color="1C2541", fill_type="solid")
            
        for _, row in df.iterrows(): ws2.append([str(item) if not isinstance(item, bool) else str(item) for item in row])
        ws2.auto_filter.ref = ws2.dimensions
        ws2.freeze_panes = "A2"
        
        col_custo = None
        for i, col in enumerate(df.columns):
            if col == 'custo_atendimento': col_custo = get_column_letter(i + 1)
        
        if col_custo:
            regra_barra = DataBarRule(start_type='num', start_value=0, end_type='max', color="FF5A00", showValue="None")
            ws2.conditional_formatting.add(f'{col_custo}2:{col_custo}{ws2.max_row}', regra_barra)

        wb.save(nome_excel)

        # --- PDF BLINDADO CONTRA ERROS DE XML ---
        nome_pdf = "/tmp/Laudo_Auditoria.pdf"
        doc = SimpleDocTemplate(nome_pdf, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
        story = []
        styles = getSampleStyleSheet()
        
        st_capa_tit = ParagraphStyle('CapaTit', parent=styles['Heading1'], fontSize=26, textColor=colors.HexColor('#0B132B'), alignment=1, spaceAfter=15, fontName="Helvetica-Bold")
        st_capa_sub = ParagraphStyle('CapaSub', parent=styles['Heading2'], fontSize=14, textColor=colors.HexColor('#FF5A00'), alignment=1)
        st_tit = ParagraphStyle('Tit', parent=styles['Heading1'], fontSize=15, textColor=colors.HexColor('#0B132B'), spaceAfter=15, fontName="Helvetica-Bold")
        st_sub = ParagraphStyle('Sub', parent=styles['Heading2'], fontSize=11, textColor=colors.HexColor('#FF5A00'), spaceBefore=15, spaceAfter=8, fontName="Helvetica-Bold")
        st_txt = ParagraphStyle('Txt', parent=styles['Normal'], fontSize=9.5, textColor=colors.HexColor('#334155'), leading=14, spaceAfter=10)
        st_alerta = ParagraphStyle('Alerta', parent=styles['Normal'], fontSize=11, textColor=colors.HexColor('#DC2626'), leading=15, fontName="Helvetica-Bold")
        
        story.append(Spacer(1, 100))
        if os.path.exists(LOGO_PATH): story.append(RLImage(LOGO_PATH, width=220, height=55))
        story.append(Spacer(1, 50))
        story.append(Paragraph("RELATÓRIO DE DATA SCIENCE E RISCO", st_capa_tit))
        story.append(Paragraph("Previsão de Churn, Desperdício Operacional e Eficiência", st_capa_sub))
        story.append(Spacer(1, 150))
        story.append(Paragraph(f"Machine Learning Processing Date: {datetime.now().strftime('%d/%m/%Y')}", ParagraphStyle('C', alignment=1, textColor=colors.HexColor('#64748B'))))
        story.append(PageBreak()) 
        
        story.append(Paragraph("1. RADAR DE RISCO E ALERTAS (CDO VIEW)", st_tit))
        
        caixa_texto = [
            [Paragraph("<b>CRITICAL ALERT: RISCO JURÍDICO E CHURN</b>", ParagraphStyle('B', fontSize=10, textColor=colors.HexColor('#7F1D1D')))],
            [Paragraph(f"O algoritmo de NLP detetou <b>{qtd_clientes_risco} clientes</b> com alta probabilidade de cancelamento ou ameaça de processos legais. "
                       f"Paralelamente, o gargalo consumiu <b>R$ {prejuizo_estimado:,.2f}</b> em folha de pagamento improdutiva.", st_alerta)]
        ]
        tabela_destaque = Table(caixa_texto, colWidths=[460])
        tabela_destaque.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#FEF2F2')),
            ('BOX', (0,0), (-1,-1), 1.5, colors.HexColor('#DC2626')),
            ('LEFTMARGIN', (0,0), (-1,-1), 15), ('TOPPADDING', (0,0), (-1,-1), 10), ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ]))
        story.append(tabela_destaque)
        story.append(Spacer(1, 15))
        
        for paragrafo in texto_ia.split('\n'):
            texto = paragrafo.replace('**', '').strip()
            if not texto: continue
            
            # BLINDAGEM ATIVADA: Limpa os caracteres que quebram o PDF (<, >, &)
            texto_seguro = html.escape(texto)
            
            if texto_seguro[0].isdigit() and '.' in texto_seguro[:3]: story.append(Paragraph(texto_seguro, st_sub))
            else: story.append(Paragraph(texto_seguro, st_txt))
                
        story.append(PageBreak()) 
        
        story.append(Paragraph("2. TOPOLOGIA VISUAL DE DADOS", st_tit))
        tabela_g1 = Table([[RLImage(path_horario, width=460, height=180)]])
        story.append(tabela_g1)
        story.append(Spacer(1, 10))
        
        tabela_g2 = Table([[RLImage(path_motivos, width=230, height=140), RLImage(path_risco, width=230, height=140)]])
        story.append(tabela_g2)
        story.append(Spacer(1, 10))
        
        tabela_g3 = Table([[RLImage(path_atendentes, width=230, height=140)]])
        tabela_g3.setStyle(TableStyle([('ALIGN', (0,0), (-1,-1), 'CENTER')]))
        story.append(tabela_g3)
        
        doc.build(story, onFirstPage=adicionar_rodape, onLaterPages=adicionar_rodape)

        with open(nome_excel, "rb") as f: exc_64 = base64.b64encode(f.read()).decode('utf-8')
        with open(nome_pdf, "rb") as f: pdf_64 = base64.b64encode(f.read()).decode('utf-8')

        return jsonify({"status": "sucesso", "mensagem": "Auditoria Nível CDO Concluída!", "excel": exc_64, "pdf": pdf_64}), 200

    except Exception as e:
        # A CAIXA NEGRA: Se algo falhar, isto imprime o erro exato nos logs do Render
        traceback.print_exc()
        return jsonify({"status": "erro", "mensagem": f"Erro interno do servidor: {str(e)}"}), 500

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
